import os
import xml.etree.ElementTree as ET
import shutil
from openpyxl import load_workbook
from datetime import datetime, timedelta
import json


file_name = "Excelsheet.xlsx"
sheet_name = "XML data"

# read config.json


final_data = []
cols = []
# Load column names from Excel without pandas
wb = load_workbook(file_name)
sheet = wb[sheet_name]
for cell in sheet[1]:
    cols.append(cell.value)

template = {}
for col in cols:
    template[col] = 0

def process_xml_file(file_path):
    temp_template = template.copy()
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            xml_content = f.read()
        
        # Wrap in a fake root since the XML lacks a single root element
        wrapped_xml = f"<FakeRoot>{xml_content}</FakeRoot>"
        root = ET.fromstring(wrapped_xml)
        
        total_count = 0
        total_amount = 0

        print(f"Processing: {file_path}")
        for item in root.findall("ValueLinkLineItem"):
            data = extract_data(item)
            print(data)
            total_count += data["Count"]
            total_amount += data["Amount"]
            temp_template["Merchant Id"] = data["Merchant Number"]
            temp_template["Store Number"] = data["Alt Merchant Number"]
            temp_template[data["Category"]+" Count"] += data["Count"]
            temp_template[data["Category"]+" Amount"] += data["Amount"]
        
        # divide each amount by 100
        for key in temp_template:
            if "Amount" in key:
                temp_template[key] = int(temp_template[key] / 100)
        temp_template["Total Activity Count"] = total_count
        temp_template["Total Activity Amount"] = int(total_amount/100)
        
        final_data.append(temp_template)
    except ET.ParseError as e:
        print(f"Error parsing {file_path}: {e}")

def extract_data(item):
    amount_elem = item.find("amount/FSNDollarUS")
    marchent_number = item.find("merchantNumber").text
    alt_marchent_number = item.find("altMerchantNumber").text
    amount = amount_elem.attrib.get("amount", "N/A") if amount_elem is not None else "N/A"
    category = item.find("category").text if item.find("category") is not None else "N/A"
    count = item.find("count").text if item.find("count") is not None else "N/A"
    return {
        "Merchant Number": marchent_number,
        "Alt Merchant Number": alt_marchent_number,
        "Amount": int(amount),
        "Category": category,
        "Count": int(count)
    }

def write_to_excel(data, original_file, sheet_name, backup_file="backup.xlsx"):
    # Create a backup copy of the original file
    shutil.copy(original_file, backup_file)
    print(f"Backup created: {backup_file}")

    try:
        # Load the copied workbook
        book = load_workbook(backup_file)

        # Write data to an existing sheet without affecting others
        sheet = book.create_sheet(sheet_name) if sheet_name not in book.sheetnames else book[sheet_name]
        
        # Clear existing data in the sheet
        sheet.delete_rows(1, sheet.max_row)
        
        # Write header
        for col, value in enumerate(cols):
            sheet.cell(row=1, column=col+1).value = value
        
        # Write data
        for row, item in enumerate(data, start=2):
            for col, key in enumerate(cols):
                sheet.cell(row=row, column=col+1).value = item.get(key, "")
        
        book.save(backup_file)
        
        print(f"Data written successfully to '{sheet_name}' in {backup_file}")

    except FileNotFoundError:
        print(f"File {original_file} not found. Creating a new file...")
        # Create a new workbook and write data
        book = load_workbook()
        sheet = book.active
        sheet.title = sheet_name
        
        # Write header
        for col, value in enumerate(cols):
            sheet.cell(row=1, column=col+1).value = value
        
        # Write data
        for row, item in enumerate(data, start=2):
            for col, key in enumerate(cols):
                sheet.cell(row=row, column=col+1).value = item.get(key, "")
        
        book.save(backup_file)

def main():
    with open("config.json", "r") as f:
        config_data = json.load(f)

    date = (datetime.now() + timedelta(days=config_data["day"])).strftime('%Y%m%d')
    input_folder = config_data["input_folder"].replace("DATE", date)


    date = (datetime.now() + timedelta(days=config_data["day"])).strftime('%m.%d.%Y')

    output_file = config_data["output_file"].replace("DATE", date)
    if not os.path.exists(input_folder):
        print(f"Input folder '{input_folder}' does not exist.")
        return

    for filename in os.listdir(input_folder):
        if filename.endswith(".xml"):
            file_path = os.path.join(input_folder, filename)
            process_xml_file(file_path)

    write_to_excel(final_data, file_name, sheet_name, backup_file=f"{output_file}.xlsx")

if __name__ == "__main__":
    main()
