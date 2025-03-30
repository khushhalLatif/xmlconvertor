import sys
from openpyxl import load_workbook
import csv
import json
import zipfile
import io
# read config.json
with open("config.json", "r") as f:
    config_data = json.load(f)

data_file = config_data["data_file"]
sheet_name = "Holding Pen View"


mapping = {
    1:"Edit Delete",
    2:25,
    3:8,
    4:11,
    5:23,
    6:16,
    7:10,
    8:9,
    9:"Tran Description",
    10:15,
    11:18,
    12:17,
    13:14,
    14:19,
    15:26,
}

# sort mapping on keys
mapping = {k: v for k, v in sorted(mapping.items(), key=lambda item: item[0])}



def write_to_excel(data, file_name, sheet_name):

    try:
        # Load the copied workbook
        book = load_workbook(file_name)

        # Write data to an existing sheet without affecting others
        sheet = book.create_sheet(sheet_name) if sheet_name not in book.sheetnames else book[sheet_name]
        
        # Clear existing data in the sheet
        sheet.delete_rows(11, sheet.max_row)

        
        # Write data
        for row, item in enumerate(data, start=11):
            for col, key in enumerate(item):
                # print(col)
                sheet.cell(row=row, column=col+1).value = key
        
        book.save(file_name)
        
        print(f"Data written successfully to '{sheet_name}' sheet in '{file_name}'")

    except Exception as e:
        _, _, tb = sys.exc_info()
        print(f"An error occurred: {e}")
        print(f"Traceback: {tb.tb_lineno}")

def load_data(data_file):
    # data file is zip
    final_data = []
    with zipfile.ZipFile(data_file) as z:
        file = z.namelist()[0]
    # load csv data into row of rows
        with z.open(file) as f:
            text_file = io.TextIOWrapper(f, encoding="utf-8")  # Convert binary to text
            reader = csv.reader(text_file)
            data = list(reader)
    
    for row_index,row in enumerate(data):
        temp_row = []
        # for col_index,col in enumerate(row):
        for key, value in mapping.items():
            # check if value is number
            if isinstance(value, int):
                temp_row.append(row[value-1])
            else:
                temp_row.append(value)
        final_data.append(temp_row.copy())


    return final_data

def main():
    file_name = "template.xlsx"
    mapped_data = load_data(data_file)
    print(mapped_data)

    write_to_excel(mapped_data, file_name, sheet_name)

if __name__ == "__main__":
    main()
